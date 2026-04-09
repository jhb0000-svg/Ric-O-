import docx
import openpyxl

def parse_docx(file_path: str) -> str:
    """Reads a DOCX file and returns its textual content."""
    doc = docx.Document(file_path)
    text = []
    for para in doc.paragraphs:
        if para.text.strip():
            text.append(para.text.strip())
    return "\n".join(text)

def parse_excel(file_path: str) -> str:
    """Reads an Excel file and returns its textual content."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            row_text = " ".join([str(cell) for cell in row if cell is not None])
            if row_text.strip():
                text.append(row_text.strip())
    return "\n".join(text)
